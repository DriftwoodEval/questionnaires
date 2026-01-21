import json
import re
from collections import defaultdict
from datetime import date, timedelta
from pathlib import Path
from typing import Any

import inquirer
import pandas as pd
from loguru import logger
from openpyxl.styles import Alignment, Font

from utils.custom_types import Appointment, Config
from utils.database import get_all_evaluators_info, get_appointments
from utils.google import create_gmail_draft, get_punch_list, upload_file_to_drive
from utils.misc import load_config

TRACKING_FILE = Path("piecework_output") / "reports_tracking.json"


def load_tracked_reports() -> dict[str, str]:
    """Load the set of previously tracked client IDs from disk."""
    if not TRACKING_FILE.exists():
        return {}

    try:
        with open(TRACKING_FILE) as f:
            data = json.load(f)
            client_ids = data.get("client_ids", {})
            if not isinstance(client_ids, dict):
                logger.warning(
                    "Tracked data file structure is invalid (not a dict). Resetting."
                )
                return {}
            return client_ids
    except (FileNotFoundError, json.JSONDecodeError):
        logger.exception("Failed to load previously tracked client IDs, resetting.")
        return {}


def save_tracked_reports(client_ids: dict[str, str]):
    """Save the set of tracked client IDs to disk."""
    TRACKING_FILE.parent.mkdir(parents=True, exist_ok=True)
    try:
        with open(TRACKING_FILE, "w") as f:
            json.dump({"client_ids": client_ids}, f, indent=2)
        logger.info(f"Saved {len(client_ids)} entries to {TRACKING_FILE}")
    except Exception:
        logger.exception(f"Failed to save tracking file to {TRACKING_FILE}")


def extract_writer_initials(assigned_to: str) -> str:
    """Extract only letters from the assigned to column."""
    if pd.isna(assigned_to) or not assigned_to:
        return ""
    return re.sub(r"[^a-zA-Z]", "", assigned_to)


def get_report_clients(config: Config) -> pd.DataFrame | None:
    """Find clients who have reports done, and who either: haven't been ran before, or were ran on the same day."""
    punch_list = get_punch_list(config)

    if punch_list is None:
        logger.critical("Punch list is empty")
        return None

    report_done = punch_list[
        (punch_list["Billed?"] == "TRUE")
        & (punch_list["AJP Review Done/Hold for payroll"] != "TRUE")
        & (
            (punch_list["For"] != "ADHD")
            | (
                (punch_list["For"] == "ADHD")
                & (punch_list["Evaluator"].str.lower() == "ap")
            )
        )
    ].copy()

    if report_done.empty:
        logger.info("No clients found that have reports done")
        return None

    tracked_reports = load_tracked_reports()
    today_str = date.today().strftime("%Y-%m-%d")

    logger.info(f"Loaded report history: {len(tracked_reports)}")

    new_reports = report_done[
        report_done["Client ID"].apply(
            lambda cid: cid not in tracked_reports
            or tracked_reports.get(cid) == today_str
        )
    ].copy()

    if new_reports.empty:
        logger.info("No new reports found")
        return None

    for client_id in new_reports["Client ID"]:
        if client_id not in tracked_reports:
            tracked_reports[client_id] = today_str

    save_tracked_reports(tracked_reports)

    result = new_reports[
        [
            "Client Name",
            "Client ID",
            "Assigned to OR added to report writing folder",
        ]
    ].rename(columns={"Assigned to OR added to report writing folder": "Assigned To"})

    rows_to_drop = []
    for idx, row in result.iterrows():
        assigned_to = row["Assigned To"]
        if pd.isna(assigned_to) or not str(assigned_to).strip():
            logger.error(
                f"'{row['Client Name'].strip()}' (ID: {row['Client ID'].strip()}) has empty 'Assigned To' field"
            )
            rows_to_drop.append(idx)

    if rows_to_drop:
        result = result.drop(rows_to_drop)
        logger.error(f"Dropped {len(rows_to_drop)} client(s) with no report writer")

    if result.empty:
        logger.error("No valid reports found after filtering out unassigned clients")
        return None

    result["Initials"] = result["Assigned To"].apply(extract_writer_initials)
    result["Writer Name"] = result["Initials"].apply(
        lambda initials: config.piecework.get_full_name(initials) if initials else ""
    )

    result = result.drop(columns=["Initials", "Assigned To"])

    logger.info(f"Found {len(result)} new reports")

    return result


def get_date_range() -> tuple[date, date] | None:
    """Prompt the user to select a date range (last week or week before)."""
    today = date.today()
    days_since_last_sunday = (today.weekday() + 1) % 7
    most_recent_sunday = today - timedelta(days=days_since_last_sunday)
    last_full_week_sunday = most_recent_sunday - timedelta(days=7)
    last_full_week_saturday = last_full_week_sunday + timedelta(days=6)

    week_before_sunday = last_full_week_sunday - timedelta(days=7)
    week_before_saturday = week_before_sunday + timedelta(days=6)

    questions = [
        inquirer.List(
            "date_range",
            message="Select the date range for the report",
            choices=[
                (
                    f"Last week ({last_full_week_sunday.strftime('%m-%d')} to {last_full_week_saturday.strftime('%m-%d')})",
                    (last_full_week_sunday, last_full_week_saturday),
                ),
                (
                    f"Week before last ({week_before_sunday.strftime('%m-%d')} to {week_before_saturday.strftime('%m-%d')})",
                    (week_before_sunday, week_before_saturday),
                ),
            ],
        ),
    ]
    answers = inquirer.prompt(questions)
    if answers:
        date_range = answers["date_range"]
        logger.info(f"Selected range: {date_range[0]} to {date_range[1]}")
        return date_range
    else:
        logger.info("No date range selected. Exiting.")
        return None


def get_work_counts(
    appointments: list[Appointment],
    evaluators: dict[int, dict],
    report_clients: pd.DataFrame | None = None,
) -> dict[str, dict[str, int]]:
    """Aggregates appointment counts by evaluator by type.

    Returns a dict mapping evaluator name to appointment type counts.
    """
    counts: dict[int | str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    worker_names: dict[int | str, str] = {}

    for appointment in appointments:
        if appointment.get("cancelled"):
            continue

        npi = appointment.get("evaluatorNpi")
        da_eval = appointment.get("daEval")

        if not npi or not da_eval:
            logger.warning(
                f"Skipping appointment with missing NPI or daEval: {appointment.get('id', 'N/A')}"
            )
            continue

        evaluator_info = evaluators.get(npi)
        evaluator_name = (
            evaluator_info.get("providerName")
            if evaluator_info and evaluator_info.get("providerName")
            else f"Unknown Evaluator (NPI: {npi})"
        )

        worker_names[npi] = str(evaluator_name)
        counts[npi][str(da_eval)] += 1

    if report_clients is not None and not report_clients.empty:
        for _, row in report_clients.iterrows():
            writer_name = row.get("Writer Name", "")
            if writer_name:
                existing_npi = next(
                    (npi for npi, name in worker_names.items() if name == writer_name),
                    None,
                )

                if existing_npi is not None:
                    counts[existing_npi]["REPORT"] = (
                        counts[existing_npi].get("REPORT", 0) + 1
                    )
                else:
                    worker_names[writer_name] = writer_name
                    counts[writer_name]["REPORT"] = (
                        counts[writer_name].get("REPORT", 0) + 1
                    )

    total_work_count = sum(sum(counts.values()) for counts in counts.values())
    logger.info(f"Total work piece count: {total_work_count}")

    return {
        worker_names.get(key, f"Unknown Worker (Key: {key})"): dict(da_eval_counts)
        for key, da_eval_counts in counts.items()
    }


def prepare_summary_data(
    work_counts: dict[str, dict[str, int]], config: Config
) -> list[dict]:
    """Prepares the aggregated appointment counts for the Summary Counts DataFrame.

    The evaluator name appears on its own row, followed by rows with type/count data.
    """
    summary_rows = []

    for worker_name, app_counts in work_counts.items():
        # Add name row with no type/count
        summary_rows.append(
            {
                "NAME": worker_name,
                "TYPE": "",
                "COUNT": "",
                "UNIT": "",
                "COST": "",
                "TOTAL PAY": "",
            }
        )

        # Add type/count rows with no name
        sorted_app_types = sorted(app_counts.keys())
        evaluator_total = 0.00

        for app_type in sorted_app_types:
            count = app_counts[app_type]
            unit_cost = config.piecework.get_unit_cost(worker_name, app_type)
            total_cost = count * unit_cost
            evaluator_total += total_cost
            summary_rows.append(
                {
                    "NAME": "",
                    "TYPE": app_type,
                    "COUNT": count,
                    "UNIT": unit_cost,
                    "COST": total_cost,
                    "TOTAL PAY": "",
                }
            )

        summary_rows.append(
            {
                "NAME": "",
                "TYPE": "",
                "COUNT": "",
                "UNIT": "",
                "COST": "",
                "TOTAL PAY": evaluator_total,
            }
        )

        summary_rows.append(
            {
                "NAME": "",
                "TYPE": "",
                "COUNT": "",
                "UNIT": "",
                "COST": "",
                "TOTAL PAY": "",
            }
        )

    return summary_rows


def prepare_detail_data(
    appointments: list[Appointment],
    evaluators: dict[int, dict],
    config: Config,
    report_clients: pd.DataFrame | None,
) -> dict[str, list[dict[str, Any]]]:
    """Prepares detail data, returning a dictionary mapping worker names to their detail rows."""
    worker_details: dict[str, list[dict[str, Any]]] = defaultdict(list)

    for appointment in appointments:
        if appointment.get("cancelled"):
            continue

        npi = appointment.get("evaluatorNpi")
        da_eval = appointment.get("daEval")

        if npi and da_eval:
            evaluator_info = evaluators.get(npi)
            evaluator_name = (
                evaluator_info["providerName"]
                if evaluator_info and evaluator_info.get("providerName")
                else f"Unknown Evaluator (NPI: {npi})"
            )

            row = {
                "WORKER": evaluator_name,
                "CLIENT": appointment.get("clientName", "N/A"),
                "START TIME": appointment.get("startTime").strftime(
                    "%Y-%m-%d %I:%M %p"
                ),
                "TYPE": str(da_eval),
            }
            worker_details[evaluator_name].append(row)

    if report_clients is not None and not report_clients.empty:
        for _, row in report_clients.iterrows():
            writer_name = row.get("Writer Name", "")
            client_name = row.get("Client Name", "")
            if writer_name:
                row = {
                    "WORKER": writer_name,
                    "CLIENT": client_name,
                    "START TIME": "N/A",
                    "TYPE": "REPORT",
                }
                worker_details[writer_name].append(row)

    combined_detail_rows = []

    for worker, details_list in worker_details.items():
        sorted_details = sorted(
            details_list, key=lambda k: (k["WORKER"], k["START TIME"])
        )
        worker_details[worker] = sorted_details
        combined_detail_rows.extend(sorted_details)

    worker_details["__COMBINED_DETAIL_DATA__"] = combined_detail_rows

    return worker_details


def generate_main_report(
    summary_data: list[dict],
    detail_data: list[dict],
    start_date: date,
    end_date: date,
    config: Config,
):
    """Generates a single Excel file with two sheets: Summary and Detail."""
    piecework_output_folder = Path("piecework_output")
    piecework_output_folder.mkdir(parents=True, exist_ok=True)

    filename = (
        piecework_output_folder
        / f"piecework_{start_date.strftime('%y-%m-%d')}_{end_date.strftime('%y-%m-%d')}.xlsx"
    )

    df_summary = pd.DataFrame(summary_data)
    df_detail = pd.DataFrame(detail_data)

    file_generated = False

    try:
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            df_summary.to_excel(writer, sheet_name="Summary Counts", index=False)
            df_detail.to_excel(writer, sheet_name="Details", index=False)

            workbook = writer.book
            summary_sheet = writer.sheets["Summary Counts"]

            currency_format = '"$"#,##0.00'

            for row_idx, row in enumerate(
                summary_data, start=2
            ):  # start=2 to skip header
                if row.get("UNIT"):
                    summary_sheet.cell(
                        row=row_idx, column=4
                    ).number_format = currency_format

                if row.get("COST"):
                    summary_sheet.cell(
                        row=row_idx, column=5
                    ).number_format = currency_format

                if row.get("TOTAL PAY"):
                    summary_sheet.cell(
                        row=row_idx, column=6
                    ).number_format = currency_format

            # Find the last row with data
            last_row = len(df_summary) + 1  # +1 for header row

            summary_sheet.cell(row=last_row + 2, column=1, value="GRAND TOTAL")

            formula = f"=SUM(F2:F{last_row})"

            grand_total_cell = summary_sheet.cell(row=last_row + 2, column=6)
            grand_total_cell.value = formula
            grand_total_cell.number_format = currency_format
            grand_total_cell.font = Font(bold=True)
            grand_total_cell.alignment = Alignment(horizontal="right")

            label_cell = summary_sheet.cell(row=last_row + 2, column=1)
            label_cell.font = Font(bold=True)

            # Auto-fit columns for Summary Counts sheet
            for column in summary_sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length

                adjusted_width = min(max_length + 5, 50)  # Add padding, cap at 50
                summary_sheet.column_dimensions[column_letter].width = adjusted_width

            # Auto-fit columns for Details sheet
            detail_sheet = writer.sheets["Details"]
            for column in detail_sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length

                adjusted_width = min(max_length + 5, 50)  # Add padding, cap at 50
                detail_sheet.column_dimensions[column_letter].width = adjusted_width

        logger.success(f"Wrote Excel file: {filename}")
        file_generated = True

    except Exception:
        logger.exception(f"An error occurred while writing the Excel file {filename}.")

    if file_generated:
        upload_file_to_drive(filename, config.payroll_folder_id)


def generate_individual_detail_reports(
    worker_details: dict[str, list[dict[str, Any]]],
    start_date: date,
    end_date: date,
    config: Config,
    evaluators: dict[int, dict],
):
    """Generates a separate Excel file for each worker's detail data."""
    base_output_folder = Path("piecework_output")
    worker_details.pop("__COMBINED_DETAIL_DATA__")

    for worker_name, detail_data in worker_details.items():
        if not detail_data:
            continue

        safe_worker_name = re.sub(r'[\\/:*?"<>|]', "", worker_name).strip()
        worker_folder = base_output_folder / safe_worker_name
        worker_folder.mkdir(parents=True, exist_ok=True)
        filename = (
            worker_folder
            / f"{safe_worker_name}_{start_date.strftime('%y-%m-%d')}_{end_date.strftime('%y-%m-%d')}.xlsx"
        )
        df_detail = pd.DataFrame(detail_data)

        try:
            with pd.ExcelWriter(filename, engine="openpyxl") as writer:
                df_detail.to_excel(writer, sheet_name="Details", index=False)

                detail_sheet = writer.sheets["Details"]
                for column in detail_sheet.columns:
                    max_length = max(
                        (len(str(cell.value)) for cell in column if cell.value),
                        default=0,
                    )
                    adjusted_width = min(max_length + 5, 50)
                    detail_sheet.column_dimensions[
                        column[0].column_letter
                    ].width = adjusted_width

            logger.info(f"Wrote individual detail file locally for: {worker_name}")
            link = upload_file_to_drive(
                filename, config.payroll_folder_id, safe_worker_name
            )

            if not link:
                logger.error(f"Failed to upload {filename} to Google Drive.")
                continue

            worker_email = None

            for _, evaluator_info in evaluators.items():
                if evaluator_info.get("providerName") == worker_name:
                    worker_email = evaluator_info.get("email")
                    break

            worker_email = config.piecework.payroll_emails.get(
                worker_name, worker_email
            )

            if not worker_email:
                logger.warning(
                    f"Could not find email for {worker_name}. Skipping Gmail draft."
                )
                continue

            subject = f"Pay Spreadsheet for {start_date.strftime('%m-%d-%Y')} to {end_date.strftime('%m-%d-%Y')}"
            message_text = f"Please refer to the following file to reconcile work completed and pay. Please reach out if you find any discrepancies.\n\n{link}"
            create_gmail_draft(
                subject=subject,
                to_addr=worker_email,
                message_text=message_text,
            )

        except Exception:
            logger.exception(
                f"An error occurred while writing the individual Excel file for {worker_name}."
            )


def main():
    """Main function to run piecework."""
    logger.info("Starting...")

    _, config = load_config()
    date_range = get_date_range()

    if not date_range:
        return

    start_date, end_date = date_range

    evaluators = get_all_evaluators_info(config)
    if not evaluators:
        logger.error("Could not load evaluator information. Aborting.")
        return

    appointments = get_appointments(config, start_date, end_date)
    if not appointments:
        logger.info("No appointments found for the selected range.")
        appointments = []

    try:
        report_clients = get_report_clients(config)
    except Exception:
        logger.exception("Failed to fetch report clients.")
        report_clients = None

    if not appointments and (report_clients is None or report_clients.empty):
        logger.info("No appointments or report writing entries found.")
        return

    appointment_counts = get_work_counts(appointments, evaluators, report_clients)
    summary_data = prepare_summary_data(appointment_counts, config)
    worker_details = prepare_detail_data(
        appointments, evaluators, config, report_clients
    )

    combined_detail_data = worker_details.get("__COMBINED_DETAIL_DATA__", [])

    if not summary_data and not combined_detail_data:
        logger.info("No valid appointments found to include in the report.")
        return

    generate_main_report(
        summary_data, combined_detail_data, start_date, end_date, config
    )
    generate_individual_detail_reports(
        worker_details, start_date, end_date, config, evaluators
    )


if __name__ == "__main__":
    main()
